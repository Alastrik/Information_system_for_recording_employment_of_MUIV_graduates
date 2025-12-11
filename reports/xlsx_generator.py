# reports/xlsx_generator.py
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import DatabaseConnection
from datetime import datetime
import os

class XlsxReportGenerator:
    @staticmethod
    def generate_employment_report(output_path):
        """Генерирует .xlsx отчёт на основе данных из БД."""
        db = DatabaseConnection()
        db.connect()

        try:
            query = """
                SELECT 
                    g.full_name,
                    g.graduation_year,
                    f.name AS faculty,
                    c.name AS company,
                    p.title AS position,
                    es.status_name AS status,
                    e.salary,
                    e.start_date,
                    e.is_current
                FROM graduate g
                LEFT JOIN user_profile up ON g.user_profile_id = up.id
                LEFT JOIN faculty f ON up.faculty_id = f.id
                LEFT JOIN employment e ON g.id = e.graduate_id
                LEFT JOIN company c ON e.company_id = c.id
                LEFT JOIN position p ON e.position_id = p.id
                LEFT JOIN employment_status es ON e.status_id = es.id
                ORDER BY g.full_name, e.start_date DESC
            """
            data = db.execute_query(query, fetch=True)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Трудоустройство"

            # Заголовок
            ws.merge_cells('A1:I1')
            ws['A1'] = 'Отчёт о трудоустройстве выпускников МУИВ'
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal="center")

            ws.merge_cells('A2:I2')
            ws['A2'] = f'Сформировано: {datetime.now().strftime("%d.%m.%Y %H:%M")}'
            ws['A2'].font = Font(size=10, italic=True)
            ws['A2'].alignment = Alignment(horizontal="center")

            ws.append([])  # пустая строка

            # Заголовки таблицы
            headers = ["ФИО", "Год выпуска", "Факультет", "Организация", "Должность", "Статус", "Зарплата (руб.)", "Дата начала", "Текущее"]
            ws.append(headers)
            header_font = Font(bold=True)
            for col in range(1, 10):
                ws.cell(row=4, column=col).font = header_font
                ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")

            # Данные
            for row in data:
                ws.append([
                    row[0] or "—",
                    row[1] or "—",
                    row[2] or "—",
                    row[3] or "—",
                    row[4] or "—",
                    row[5] or "—",
                    f"{row[6]:,.2f}" if row[6] else "—",
                    row[7].strftime("%d.%m.%Y") if row[7] else "—",
                    "Да" if row[8] else "Нет"
                ])

            # Автоподбор ширины
            for col in range(1, 10):
                max_length = 0
                column = get_column_letter(col)
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column].width = adjusted_width

            # Сохранение
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            return output_path

        finally:
            db.disconnect()